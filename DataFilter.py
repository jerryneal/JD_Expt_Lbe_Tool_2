
''' 
    Tilte       : DataFilter.py
    Author      : Neal Bazirake
    Description : To correctly clean data and filter out data along with calculations to systematically remove unecessarily addons.
    Date        : 1/20/2015
'''

import os,pyodbc
import openpyxl
from openpyxl import Workbook,load_workbook #allows manipulation of excel workbook / worksheet
from datetime import datetime, timedelta
sheets = ['2014Q4','2014Q3','2014Q2','2014Q1','2013Q4','2013Q3','2013Q2','2013Q1','2012Q4']
clean_outs = ['Serial Number','=============','TABLE 1','Engine','TBD','N/A','To be provided after build','tbd'
              'Not Used','Serial Number','Date','Added','Removed','Not Assigned','Not Used','tbd']


class DataFilter():

    def __init__(self, strdate, endate):
      self.strdate = strdate
      self.endate  = endate


    def datapuller(self):
      '''
       #Getting historical data
       '''
       os.chdir("C:/Users/nb20805/Desktop/Testbase")
       wb = load_workbook('Historical_data.xlsx', use_iterators = True)
       previous_quarter_adds,previous_quarter_close =[],[]
       for i in sheets:
           ws = wb.get_sheet_by_name(''+i)
          
           for row in ws.iter_rows():
                if row[0].column == 'A' and row[0].internal_value != None:
                     previous_quarter_adds.append(row[0].internal_value)
                if row[0].column == 'D' and row[0].internal_value != None:
                     previous_quarter_close.append(row[0].internal_value)
         
           #print('/n****End of quarter break ****/n')
       
       #Below are lists containing the current opens and closes
       
       self.previous_quarter_adds = list(set(previous_quarter_adds) - set(clean_outs))
       self.previous_quarter_close = list(set(previous_quarter_close) - set(clean_outs))
       
       '''
       #Pulling current quarter adds (These are 210 pulls)
       '''

       self.strdate = "2013-01-01" #input("Enter the startdate:")
       self.endate = "2013-03-30" #input("Enter the end date:")
       self.strdate = datetime.strptime(self.strdate, "%Y-%m-%d")
       self.endate = datetime.strptime(self.endate, "%Y-%m-%d")

       db = pyodbc.connect(driver = '{SQL Server}', server ='############', database = '#########DB')
       cursor = db.cursor()

       self.engines_in_quarterly_add = [] ##Container to store our data

       results_01 = cursor.execute("""SELECT engine_serial_no, esn_entered_date 
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE label_choice = '1068.210 - Testing Exemption'
                                    AND esn_entered_date BETWEEN ? AND ?
                                    AND exempt_number > 7000
                                    AND family_choice IN ('NONROAD','LSI','CNG',NULL)
                                    AND country IN('USA','USA/Outside USA', NULL)
                                    AND current_status = 'Active'
                                    """,self.strdate,self.endate)
           
       rows_01 = results_01.fetchall()
       for item in rows_01:
            self.engines_in_quarterly_add.append(item[0])

       #print(engines_in_quarterly_add) ##This is a quick checker to ensure that the actual data is being pulled.

       '''     
       #Pullling current quarter removes (These are 210 pulls)
       '''     
       self.engines_in_quarterly_remove = []
       results_02 = cursor.execute("""SELECT engine_serial_no, date_completed 
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE label_choice = '1068.210 - Testing Exemption'
                                    AND date_completed BETWEEN ? AND ?
                                    AND exempt_number > 7000
                                    AND family_choice IN ('NONROAD','LSI','CNG',NULL)
                                    AND country IN('USA','USA/Outside USA', NULL)
                                    """,self.strdate,self.endate)
           
       rows_02 = results_02.fetchall()
       for item in rows_02:
            self.engines_in_quarterly_remove.append(item[0])

       #print(engines_in_quarterly_remove) ##This is a quick checker to ensure that the actual data is being pulled.
       '''     
       #Pullling all 215's to 210's - These are part of addins
             1.  Get all 210s closeouts (previous closeouts)
             2.  Get all 215 adds for this quarter
             3.  Subtract and get the difference
       '''

       self.engine_215_adds = []
       results_03 = cursor.execute("""SELECT engine_serial_no 
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE label_choice <> '1068.210 - Testing Exemption'
                                    AND esn_entered_date BETWEEN ? AND ?
                                    AND exempt_number > 7000
                                    AND family_choice IN ('NONROAD','LSI','CNG',NULL)
                                    AND country IN('USA','USA/Outside USA', NULL)
                                    """,self.strdate,self.endate)
           
       rows_03 = results_03.fetchall()
       for item in rows_03:
            self.engine_215_adds.append(item[0])

       return self.engines_in_quarterly_add, self.engine_215_adds.append


       #print(engine_215_adds)
       other_Addins = list(set(previous_quarter_close) - set(engine_215_adds))
       other_Addins = list(set(other_Addins) - set(clean_outs))
       #print(other_Addins)

       
       
       '''     
       #Pullling all 210's to 215's - These are part of closeouts
            1. Get all 215s closeouts
            2. Get all 210 adds for this quarter
            3. Subtract and get the difference of these closeouts.
       '''

       self.engines_in_quarterly_remove_2 = []
       results_03 = cursor.execute("""SELECT engine_serial_no, date_completed 
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE label_choice <> '1068.210 - Testing Exemption'
                                    AND date_completed BETWEEN ? AND ? 
                                    AND exempt_number > 7000
                                    AND family_choice IN ('NONROAD','LSI','CNG',NULL)
                                    AND country IN('USA','USA/Outside USA', NULL)
                                    """,self.strdate,self.endate)
           
       rows_03 = results_03.fetchall()
       for item in rows_03:
            self.engines_in_quarterly_remove_2.append(item[0])
       
       other_Closes = list(set(engines_in_quarterly_remove_2) - set(previous_quarter_adds))
       other_Closes = list(set(other_Closes) - set(clean_outs))
       #print(other_Closes)


       #----------------------------------------------#------------------------------------------
       #Concatenate the adds on one side and removes

       #Adds
       temp_add_column_engines = list(set(engines_in_quarterly_add) - set(previous_quarter_adds)) + other_Addins

       #Removes
       temp_remove_column_engines = list(set(engines_in_quarterly_remove) - set(previous_quarter_close)) + other_Closes
       temp_remove_column_engines = list(set(temp_remove_column_engines) - set(clean_outs))
       temp_remove_colum_engines = list(set(temp_remove_column_engines) - set(previous_quarter_close))
       
       #--------------------------------------#----------------------------------------------------

    def writeToexcel(self):
       '''
       Send items to get other information and send to excel file
       '''
       os.chdir("C:/Users/#####/Desktop/Testbase")
       wb = load_workbook('quarterly_output.xlsx')
       ws = wb.create_sheet()
       ws.title = "New Quarterly"
       ws.cell('A1').value = "Engine Serial Number"
       ws.cell('B1').value = "Date Added"
       ws.cell('D1').value = "Engine Serial Number"
       ws.cell('E1').value = "Date Added"
       ws.cell('F1').value = "Date Removed"
       
       
       row_num = 4
       for i in temp_add_column_engines:
            results_03 = cursor.execute("""SELECT engine_serial_no, esn_entered_date 
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE engine_serial_no = ?
                                    """, i)
            rw1 = results_03.fetchone()
            #print(rw1)
            ws.cell("A"+str(row_num)).value = rw1[0]
            ws.cell("B"+str(row_num)).value = rw1[1]
            row_num +=1
       row_num = 4     
       for i in temp_remove_column_engines:
            results_04 = cursor.execute("""SELECT engine_serial_no, esn_entered_date , date_completed
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE engine_serial_no = ?
                                    """, i)
            rw2 = results_04.fetchone()
            #print(rw2)
            ws.cell("D"+str(row_num)).value = rw2[0]
            ws.cell("E"+str(row_num)).value = rw2[1]
            ws.cell("F"+str(row_num)).value = rw2[2]
            row_num+=1
      
       return wb.save("quarterly_output.xlsx")



    
