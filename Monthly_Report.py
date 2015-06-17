'''
  Title   : Monthly_Report.py
  Author  :   Neal Bazirake
  Date    :   03/01/2014
  Description   : The Month class includes all sub-SQL querries as well as the writing the results into excel sheets
'''


import pyodbc # Link to get SQL commands
import time   # Gives us time functionality
from datetime import datetime, timedelta # Additional and more advanced time functionality
import csv     #Writing files over to comma seperated value sheets in excel
import openpyxl # Directly editing files into .xls and .xlsx files
from openpyxl import Workbook # Interfacing with microsoft excel workbooks
from openpyxl import load_workbook #allows manipulation of excel workbook / worksheet
from openpyxl.style import Color, Fill, Alignment # Adding color mechanisms and designs to the excel file to suite user requirements
import os   # To access paths into primary memory.
from win32com.client import Dispatch #Python-Windows client to use Outlook functions like sending emails after getting results.

'''
This makes the Monthly report in a Brute-force manner iterating the database for each cell of the excel spreadsheet
'''

server = '##SERVERNAME'
database = '##DATABASENAME'
class Monthly:
    
    def __init__(self):
        pass
        
    def monthly_queries(self,strdate,enddate):

         #New Spreadsheet
         os.chdir("//#######/####_####_restricted/Exemption quarterly reports/2014")
         wb = load_workbook('montly_report.xlsx')
         ws = wb.create_sheet()
         ws.title = "New Monthly"
        
         #1068.210 - Active -  Total Active - CI/LSI(US) - in B5
         db = pyodbc.connect(driver = '{SQL Server}', server ='##SERVERNAME', database = '##DATABASENAME')
         cursor = db.cursor()
         results1 = cursor.execute("""SELECT exempt_number,eng###_serial_no
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > '7000'
                                   AND current_status = 'Active'
                                   AND family_choice != '#####'
                                   AND country != 'Outside USA'
                                   AND label_choice != '###8.@#5 - Manufacturer Owned'
                                   """)
         rows1 = results1.fetchall()
         for item in rows1:
             self.index1 = rows1.index(item)
             self.index1 += 1
         ws.cell('B5').value = self.index1
         
         #1068.210 - Active -  Total Active - CI/LSI(nonUS) - in B6
         results2 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND country = 'Outside USA'
                                   AND current_status = 'Active'
                                   AND label_choice <> '###8.@#5 - Manufacturer Owned'
                                   """)
         rows2 = results2.fetchall()
         for item in rows2:
             self.index2 = rows2.index(item)
             self.index2 += 1
         ws.cell('B6').value = self.index2

         
         #1068.210 - Active - Legacy Labels - in B7
         results_l = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND family_choice <> '##R###'
                                   AND current_status = 'Active'
                                   AND label_choice <> '###8.@#5 - Manufacturer Owned'
                                   """)
         rows2_1 = results_l.fetchall()
         for item in rows2_1:
             self.index2_1 = rows2_1.index(item)
             self.index2_1 += 1
         ws.cell('B7').value = self.index2_1

         #1068.210 - Active -  Newly Issued - CI/LSI - in C5
         results3 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS
                                   WHERE label_choice <> '1#### - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND label_needed_date BETWEEN ? AND ?
                                   AND country IN ('USA','USA/O####e USA')
                                   """, strdate, enddate)
         
         rows3 = results3.fetchall()
         self.index3 = 0
         for item in rows3:
             self.index3 = rows3.index(item)
             self.index3 += 1
         ws.cell('C5').value = self.index3


         #1068.210 - Active -  Newly Issued - CI/LSI(nonUS) - in C6
         results5 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS
                                   WHERE label_choice = '1068.210 - Testing Exemption'
                                   AND exempt_number > (7000)
                                   AND label_needed_date BETWEEN ? AND ?
                                   AND country = 'Outside USA'
                                   """, strdate,enddate)
         rows5 = results5.fetchall()
         self.index5 = 0
         for item in rows5:
             self.index5 = rows5.index(item)
             self.index5 += 1
         ws.cell('C6').value = self.index5

         
         #1068.210 - Active -  Updated in Last 3 months( by Admin+user) - CI/LSI(US) - in D5
         results6 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND country <> 'Outside USA'
                                   AND current_status IN ('Active','Stored in Warehouse')
                                   AND last_updated_date BETWEEN ? AND ?
                                   AND label_choice <> '###8.@#5 - Manufacturer Owned'
                                   """, strdate,enddate)
         rows6 = results6.fetchall()
         self.index6 = 0
         for item in rows6:
             self.index6 = rows6.index(item)
             self.index6 += 1
         ws.cell('D5').value = self.index6
     

         #1068.210 - Active -  Updated in Last 3 months( by Admin+user) - CI/LSI(nonUS) - in D6
         results7 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND country = 'Outside USA'
                                   AND current_status = 'Active'
                                   AND last_updated_date BETWEEN ? AND ?
                                   AND label_choice <> '###8.@#5 - Manufacturer Owned'
                                   """, strdate,enddate)
         rows7 = results7.fetchall()
         self.index7 = 0
         for item in rows7:
             self.index7 = rows7.index(item)
             self.index7 += 1
         ws.cell('D6').value = self.index7
       
         #1068.210 - Active -  Update past due( by Admin+user) - CI/LSI(US) - in E5
         results8 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Active'
                                   AND country <> 'Outside USA'
                                   AND label_choice <> '###8.@#5 - Manufacturer Owned'
                                   AND last_updated_date <= DATEADD(month, -3, ?)
                                   """,enddate)
         rows8 = results8.fetchall()
         self.index8 = 0
         for item in rows8:
             self.index8 = rows8.index(item)
             self.index8 += 1
         ws.cell('E5').value = self.index8


         #1068.210 - Active -  Update past due( by Admin+user) - CI/LSI(nonUS) - in E6
         results9 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND current_status = 'Active'
                                   AND country = 'Outside USA'
                                   AND last_updated_date < DATEADD(month, -3, ?)
                                   AND label_choice <> '###8.@#5 - Manufacturer Owned'
                                   """, enddate)
         rows9 = results9.fetchall()
         self.index9 = 0
         for item in rows9:
             self.index9 = rows9.index(item)
             self.index9 += 1
         ws.cell('E6').value = self.index9


         #1068.210 - Warehoused - Active - CI/LSI(nonUS) - in E6
        
         #1068.210 - Active - Labels with no ESNs - CI/LSI(US) - in G5
         results12 = cursor.execute("""SELECT exempt_number
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE eng###_serial_no NOT LIKE '[A-Z][A-Z][0-9][0-9][0-9][0-9][A-Z][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9]'
                                    AND exempt_number > (7000)
                                    AND family_choice <> '##R###'
                                    AND current_status = 'Active'
                                    AND label_choice = '1068.210 - Testing Exemption'
                                    AND country <> 'Outside USA'
                                   """)
         rows12 = results12.fetchall()
         self.index12 = 0
         for item in rows12:
             self.index12 = rows12.index(item)
             self.index12 += 1
         ws.cell('G5').value = self.index12


         #1068.210 - Active - Labels with no ESNs - CI/LSI(nonUS) - in G6
         results13 = cursor.execute("""SELECT exempt_number
                                    FROM ###_EXEMPT_LABELS 
                                    WHERE eng###_serial_no NOT LIKE '[A-Z][A-Z][0-9][0-9][0-9][0-9][A-Z][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9]'
                                    AND exempt_number > (7000)
                                    AND family_choice <> '##R###'
                                    AND current_status = 'Active'
                                    AND label_choice = '1068.210 - Testing Exemption'
                                    AND country = 'Outside USA'
                                    """)
         rows13 = results13.fetchall()
         self.index13 = 0
         for item in rows13:
             self.index13 = rows13.index(item)
             self.index13 += 1
         ws.cell('G6').value = self.index13
     

         #----------------------------------------------------------#-------------------------------------------------

         ####8.@#5 - Active -  Total Active - CI/LSI(US) - in H5
         results14 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND country <> 'Outside USA'
                                   AND current_status = 'Active'
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   OR label_choice = 'NULL'
                                   """)
         rows14 = results14.fetchall()
         self.index14 = 0
         for item in rows14:
             index14 = rows14.index(item)
             index14 += 1
         ws.cell('H5').value = index14


         ####8.@#5 - Active -  Total Active - CI/LSI(nonUS) - in H6
         results15 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND country = 'Outside USA'
                                   AND current_status = 'Active'
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   """, )
         rows15 = results15.fetchall()
         self.index15 = 0
         for item in rows15:
             self.index15 = rows15.index(item)
             self.index15 += 1
         ws.cell('H6').value = self.index15



         ####8.@#5 - Active -  Newly Issued - CI/LSI - in I5
         results16 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND country <> 'Outside USA'
                                   AND request_date BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows16 = results16.fetchall()
         self.index16 = 0
         for item in rows16:
             self.index16 = rows16.index(item)
             self.index16 += 1
         ws.cell('I5').value = self.index16
       
         
         ####8.@#5 - Active -  Newly Issued - in I6
         results17 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND country = 'Outside USA'
                                   AND label_needed_date BETWEEN ? AND ?
                                   """, enddate,enddate)
         rows17 = results17.fetchall()
         self.index17 = 0
         for item in rows17:
             self.index17 = rows17.index(item)
             self.index17 += 1
         ws.cell('I6').value = self.index17
       

         ####8.@#5 - Active -  Updated in Last 3 months( by Admin+user) - CI/LSI(nonUS) - in J5
         results18 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '##R###'
                                   AND country <> 'Outside USA'
                                   AND last_updated_date BETWEEN DATEADD(month, -3, ?) AND ?
                                   """, enddate,enddate)
         rows18 = results18.fetchall()
         self.index18 = 0
         for item in rows18:
             self.index18 = rows18.index(item)
             self.index18 += 1
         ws.cell('J5').value = self.index18


         ####8.@#5 - Active -  Update past due - CI/LSI(US) - in K5
         results19 = cursor.execute("""SELECT exempt_number
                                   FROM EPA_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND country <> 'Outside USA'
                                   AND last_updated_date < DATEADD(month, -3, ?)
                                   """, enddate)
         rows19 = results19.fetchall()
         self.index19 = 0
         for item in rows19:
             self.index19 = rows19.index(item)
             self.index19 += 1
         ws.cell('K5').value = self.index19


         ####8.@#5 - Active -  Update past due - CI/LSI(nonUS) - in K6
         results20 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND country = 'Outside USA'
                                   AND last_updated_date < DATEADD(month, -3, ?)
                                   """, enddate)
         rows20 = results20.fetchall()
         self.index20 = 0
         for item in rows20:
             self.index20 = rows20.index(item)
             self.index20 += 1
         ws.cell('K6').value = self.index20


         ####8.@#5 - Warehoused - Active - CI/LSI(nonUS) - in E6
         
         ####8.@#5 - Active - Labels with no ESNs - CI/LSI(US) - in M5
         results23 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND country <> 'Outside USA'
                                   AND eng###_serial_no NOT LIKE '[A-Z][A-Z][0-9][0-9][0-9][0-9][A-Z][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9]'
                                   """)
         rows23 = results23.fetchall()
         self.index23 = 0
         for item in rows23:
             self.index23 = rows23.index(item)
             self.index23 += 1
         ws.cell('M5').value = self.index23
      

         ####8.@#5 - Active - Labels with no ESNs - CI/LSI(nonUS) - in M6
         results24 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE current_status IN ('Active','Stored in Warehouse')
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   AND exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND country = 'Outside USA'
                                   AND eng###_serial_no NOT LIKE '[A-Z][A-Z][0-9][0-9][0-9][0-9][A-Z][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9]'
                                   """)
         rows24 = results24.fetchall()
         self.index24 = 0
         for item in rows24:
             self.index24 = rows24.index(item)
             self.index24 += 1
         ws.cell('M6').value = self.index24

         

         ####8.@#5 - Total Active - Legacy - in H7
         results25 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND current_status = 'Active'
                                   AND label_choice = '###8.@#5 - Manufacturer Owned'
                                   """)
         rows25 = results25.fetchall()
         self.index25 = 0
         for item in rows25:
             self.index25 = rows25.index(item)
             self.index25 += 1
         ws.cell('H7').value = self.index25

         

         ####8.@#5 - Newly Issued - Legacy - in H8
         self.index26 = 0
         ws.cell('I7').value = self.index26

         ####8.@#5 - Active -  Updated in Last 3 months( by Admin+user) - Legacy - in J7
         results27 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND current_status = 'Active'
                                   AND last_updated_date BETWEEN DATEADD(month, -3, ?) AND ?
                                   """, enddate,enddate)
         rows27 = results18.fetchall()
         self.index27 = 0
         for item in rows27:
             self.index27 = rows27.index(item)
             self.index27 += 1
         ws.cell('J7').value = self.index27
       

         ####8.@#5 - Active -  Update past due - CI/LSI(nonUS) - in K7
         results28 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND current_status = 'Active'
                                   AND last_updated_date <= DATEADD(month, -3, ?)
                                   """, enddate)
         rows28 = results20.fetchall()
         self.index28 = 0
         for item in rows28:
             self.index28 = rows28.index(item)
             self.index28 += 1
         ws.cell('K7').value = self.index28


         ####8.@#5 - Warehoused - Legacy - in L7
         index29 = 0
         ws.cell('L7').value = index29

         ####8.@#5 - Active - Labels with no ESNs - Legacy- in M7
         results30 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND eng###_serial_no NOT LIKE '[A-Z][A-Z][0-9][0-9][0-9][0-9][A-Z][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9][A-Z0-9]'
                                   AND current_status = 'Active'
                                   """)
         rows30 = results30.fetchall()
         self.index30 = 0
         for item in rows30:
             self.index30 = rows30.index(item)
             self.index30 += 1
         ws.cell('M7').value = self.index30
        
         #-----------------------------------------------------#--------------------------------
         #All exemption closed
         #---------------------
         #All - Closed -  Scrapped - CI/LSI(US) - in N5
         results31 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Scrapped'
                                   AND country <> 'Outside USA'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows31 = results31.fetchall()
         self.index31 = 0
         for item in rows31:
             self.index31 = rows31.index(item)
             self.index31 += 1
         ws.cell('N5').value = self.index31
       

         #All - Closed -  Scrapped - CI/LSI(nonUS) - in N6
         results32 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Scrapped'
                                   AND country = 'Outside USA'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows32 = results32.fetchall()
         self.index32 = 0
         for item in rows32:
             self.index32 = rows32.index(item)
             self.index32 += 1
         ws.cell('N6').value = self.index32
      

         #All - Closed -  Scrapped - CI/LSI(US) - in N7
         results33 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND current_status = 'Scrapped'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows33 = results33.fetchall()
         self.index33 = 0
         for item in rows33:
             self.index33 = rows33.index(item)
             self.index33 += 1
         ws.cell('N7').value = self.index33
       

         #All - Closed -  Updated and Sold - CI/LSI(US) - in O5
         results34 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Updated and Sold'
                                   AND country <> 'Outside USA'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows34 = results34.fetchall()
         self.index34 = 0
         for item in rows34:
             self.index34 = rows34.index(item)
             self.index34 += 1
         ws.cell('O5').value = self.index34
        

         #All - Closed -  Updated and Sold - CI/LSI(nonUS) - in O6
         results35 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Updated and Sold'
                                   AND country = 'Outside USA'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows35 = results35.fetchall()
         self.index35 = 0
         for item in rows35:
             self.index35 = rows35.index(item)
             self.index35 += 1
         ws.cell('O6').value = self.index35


         #All - Closed -  Updated and Sold - CI/LSI(US) - in O7
         results36 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND current_status = 'Updated and Sold'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows36 = results33.fetchall()
         self.index36 = 0
         for item in rows36:
             self.index36 = rows36.index(item)
             self.index36 += 1
         ws.cell('O7').value = self.index36
      

         #All - Closed -  Not Used - CI/LSI(US) - in P5
         results37 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Not Used'
                                   AND country <> 'Outside USA'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows37 = results37.fetchall()
         self.index37 = 0
      
         for item in rows37:
             self.index37 = rows37.index(item)
             self.index37 += 1
         ws.cell('P5').value = self.index37
        
         

         #All - Closed -  Not Used - CI/LSI(nonUS) - in P6
         results38 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number > (7000)
                                   AND family_choice <> '#####'
                                   AND current_status = 'Not Used'
                                   AND country = 'Outside USA'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows38 = results38.fetchall()
         self.index38 = 0
         for item in rows38:
             self.index38 = rows38.index(item)
             self.index38 += 1
         ws.cell('P6').value = self.index38
      

         #All - Closed -  Not Used - CI/LSI(nonUS) - in P7
         results39 = cursor.execute("""SELECT exempt_number
                                   FROM ###_EXEMPT_LABELS 
                                   WHERE exempt_number < (7000)
                                   AND current_status = 'Not Used'
                                   AND date_completed BETWEEN ? AND ?
                                   """, strdate,enddate)
         rows39 = results39.fetchall()
         self.index39 = 0

         #Function to help out in Highlighting particular cells
         def highlight_with(cellref,code):
             a1 = ws.cell(cellref)
             a1.style.fill.fill_type=openpyxl.style.Fill.FILL_SOLID
             a1.style.fill.start_color.index=code
             
         
         highlight_with('A2','FFFF00')
         highlight_with('A3','FFFF00')

         for i in ['B2','B3','C2','C3','D2','D3','E2','E3','F2','F3','G2','G3']:
             highlight_with(i,'7FFF00')
         for i in ['B4','C4','D4','E4','F4','G4']:
             highlight_with(i,'ADFF2F')
         for i in ['H2','H3','I2','I3','J2','J3','K2','K3','L2','L3','M2','M3']:
             highlight_with(i,'DAA520')
         for i in ['H4','I4','J4','K4','L4','M4']:
             highlight_with(i,'FFDEAD')
         for i in ['N2','N3','O2','O3','P2','P3']:
             highlight_with(i,'1E90FF')
         for i in ['N4','O4','P4']:
             highlight_with(i,'ADD8E6')

         #Filling in some of the spaces
         ws.cell('A5').value = "CI/LSI(US)"
         ws.cell('A6').value = "CI/LSI (non-US)"
         ws.cell('A7').value = "Legacy Labels"

         ws.cell('A3').value = "Exempt Label Report"
         ws.merge_cells('B2:G2')
         ws.cell('B2').value = "        Testing Exemptions - 1068.210"
         ws.cell('B3').value = "                  Active"
         ws.merge_cells('B3:G3')
         ws.merge_cells('H2:M2')
         ws.cell('H2').value = "          Manufacturer Owned Exemptions - ###8.@#5"
         ws.cell('H3').value = "                   Active"
         ws.merge_cells('H3:M3')

         ws.merge_cells('N2:P2')
         ws.cell('N2').value = "             All Exemption"
         ws.merge_cells('N3:P3')
         ws.cell('N3').value = "                  Closed"

         ws.cell('B4').value = "Total Active"
         ws.cell('H4').value= "Total Active"
         ws.cell('C4').value = "Newly Issued"
         ws.cell('I4').value = "Newly Issued"
         ws.cell('D4').value = "Updated in last 3 months"
         ws.cell('J4').value = "Updated in last 3 months"
         ws.cell('E4').value = "Updated Past Due"
         ws.cell('K4').value = "Updated Past Due"
         ws.cell('F4').value = "Warehoused"
         ws.cell('L4').value = "Warehoused"
         ws.cell('G4').value = "Exempt Label w/o ESN"
         ws.cell('M4').value = "Exempt Label w/o ESN"
         ws.cell('N4').value = "Scrapped"
         ws.cell('O4').value = "Updated & Sold"
         ws.cell('P4').value = "Not Used"

         #This wraps the texts into the various cells.
         for i in ['B4','H4','C4','I4','D4','J4','E4','K4','F4','L4','G4','M4','N4','O4','P4','A3','A6','A5','A7','B2'
                   ,'B3','H2','H3','N2','N3']:
             ws.cell(i).style.alignment.wrap_text = True
             
         cursor.close()
         wb.save("montly_report.xlsx")

        #This opens the excel file 
         xl = Dispatch('Excel.Application')
         wb = xl.Workbooks.Open('//#######/####_####_restricted/Exemption quarterly reports/2014/montly_report.xlsx')
         xl.Visible = True    # optional: if you want to see the spreadsheet
     
         
if __name__ == "__main__":
    Monthly()
    
   
   
   
