
'''
    Name: Weekly_generator
    Version: 1.0
    Author: Neal Bazirake
    Description: Weekly generator querried weekely results for the report generator.

    Additional Notes:
        #1. After looking through refractoring needs, there seemed not to be need for classes at this point
            but to be considered later on.
'''
import pyodbc # Link to get SQL commands
import time   # Gives us time functionality
from datetime import datetime, timedelta # Additional and more advanced time functionality
import csv     #Writing files over to comma seperated value sheets in excel
import openpyxl # Directly editing files into .xls and .xlsx files
from openpyxl import Workbook # Interfacing with microsoft excel workbooks
from openpyxl import load_workbook #allows manipulation of excel workbook / worksheet
from openpyxl.style import Color, Fill, Alignment # Adding color mechanisms and designs to the excel file to suite user requirements
import os   # To access paths into primary memory.
from win32com.client import Dispatch #Python-Windows client to use Outlook functions like sending emails after getting results.

server = '##SERVERNAME'
database = '##DATABASENAME'


def weekly_generator(strdate,enddate):

     
     os.chdir("//JSELECTED PATH")
     wb = load_workbook('Weekly_Report.xlsx')
     ws = wb.create_sheet()
     ws.title = "New Weekly "+strdate
    
     #Labels removed
     db = pyodbc.connect(driver = '{SQL Server}', server ='##SERVERNAME', database = '##DATABASENAME')
     cursor = db.cursor()
     results1 = cursor.execute("""SELECT exempt_number
                               FROM [###_EXEMPT_LABELS]
                               WHERE date_completed between ? AND ?
                               """,strdate,enddate)
     rows1 = results1.fetchall()
     for item in rows1:
         index1 = rows1.index(item)
         index1 += 1
     ws.cell('B3').value = index1
     
     #Label added
     results2 = cursor.execute("""SELECT exempt_number
                                FROM [###_EXEMPT_LABELS]
                                WHERE label_needed_date between ? AND ?
                                AND current_status = 'Active'
                                """,strdate,enddate)
     rows2 = results2.fetchall()
     for item in rows2:
         index2 = rows2.index(item)
         index2 += 1

     ws.merge_cells('B2:D2')
     ws.cell('B2').value = "Weekly report"
     ws.cell('A3').value = "Labels Added"   
     ws.cell('B3').value = index1
     ws.cell('A4').value = "Labels Removed"
     ws.cell('B4').value = index2

     #This wraps the texts into the various cells.
     for i in ['B2','A3','B3','A4','B4','B2','D2']:
             ws.cell(i).style.alignment.wrap_text = True

     Shade_cells = ['A3','A4','B2']
     for i in Shade_cells:
          a1=ws.cell(''+i)
          a1.style.fill.fill_type=openpyxl.style.Fill.FILL_SOLID
          a1.style.fill.start_color.index='F0E68C'
     
     cursor.close()
     wb.save("Weekly_Report.xlsx")
     #print(ws.title,index1,index2)
     
     #This opens the excel file 
     xl = Dispatch('Excel.Application')
     wb = xl.Workbooks.Open('//Jdshare/emis_cert_restricted/Exemption quarterly reports/2014/Weekly_Report.xlsx')
     xl.Visible = True    # optional: if you want to see the spreadsheet
     

